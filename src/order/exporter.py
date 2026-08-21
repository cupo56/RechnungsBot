"""
Excel-Export für das Order-Tab-Ergebnis (Ziel-Liste + angehängte
Order-Spalte). Werte-only, keine Übernahme der Original-Formatierung —
konsistent mit src/excel/exporter.py.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

_HEADER_FILL = PatternFill("solid", fgColor="D9D9D9")
_BOLD = Font(bold=True)


def export_order_result(headers, rows, output_path):
    """Schreibt headers/rows (siehe OrderResult) als neue .xlsx-Datei."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Order"

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = _BOLD
        cell.fill = _HEADER_FILL

    for row_idx, row in enumerate(rows, start=2):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    for col_idx, header in enumerate(headers, start=1):
        max_len = len(str(header))
        for row in rows:
            if col_idx - 1 < len(row) and row[col_idx - 1] is not None:
                max_len = max(max_len, len(str(row[col_idx - 1])))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 50)

    ws.freeze_panes = "A2"
    wb.save(output_path)

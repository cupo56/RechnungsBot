"""
Excel-Parser für RechnungsBot.
Liest die Superfeed-Excel-Datei ein und extrahiert bestellte Positionen.
"""

import re
import openpyxl

# Vorcompilierte Patterns – einmalig beim Modulimport, nicht bei jedem Aufruf
_RE_GENDER = re.compile(r'\s*\((man|woman|unisex)\)\s*', re.IGNORECASE)
_RE_COVER  = re.compile(
    r'\s*\((New Cover|Old Cover|Cover with [^)]+|Without box|[A-Za-z]+ Cover[^)]*)\)\s*',
    re.IGNORECASE,
)
_RE_WHITESPACE = re.compile(r'\s{2,}')

# Parfüm-Bezeichnungen: längere Varianten zuerst, damit kein Teilmatch vorgreift
_FRAGRANCE_REPLACEMENTS = [
    (re.compile(r'Eau De Parfum Intense', re.IGNORECASE), 'EdP Intense'),
    (re.compile(r'Eau De Parfum',         re.IGNORECASE), 'EdP'),
    (re.compile(r'Eau De Toilette',       re.IGNORECASE), 'EdT'),
    (re.compile(r'Eau De Cologne Intense',re.IGNORECASE), 'EdC Intense'),
    (re.compile(r'Eau De Cologne',        re.IGNORECASE), 'EdC'),
    (re.compile(r'Extrait de Parfum',     re.IGNORECASE), 'Extrait'),
    (re.compile(r'Parfum Intense',        re.IGNORECASE), 'Parfum Intense'),
]


def shorten_product_name(name):
    """
    Kürzt Produktbeschreibungen für die Rechnung.
    z.B. 'Boss Boss Bottled Eau De Toilette 50 ml (man)' → 'Boss Boss Bottled EdT 50 ml'
    """
    if not name:
        return ""

    text = str(name).strip()
    text = _RE_GENDER.sub(' ', text)
    text = _RE_COVER.sub(' ', text)

    for pattern, replacement in _FRAGRANCE_REPLACEMENTS:
        text = pattern.sub(replacement, text)

    # Großgeschriebene Abkürzungen normalisieren
    text = text.replace(' EDT ', ' EdT ').replace(' EDP ', ' EdP ').replace(' EDC ', ' EdC ')
    if text.endswith(' EDT'):
        text = text[:-4] + ' EdT'
    if text.endswith(' EDP'):
        text = text[:-4] + ' EdP'

    return _RE_WHITESPACE.sub(' ', text).strip()


def parse_excel(filepath):
    """
    Liest eine Superfeed-Excel-Datei und gibt bestellte Positionen zurück.

    Returns:
        list[dict]: Liste von Positionen mit Schlüsseln:
            - ean (str): EAN-Code
            - product (str): Gekürzter Produktname
            - quantity (int): Bestellmenge
            - source_price (float): Einkaufspreis aus der Quelldatei
    """
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb.active

    # Spalten automatisch erkennen anhand der Kopfzeile
    # In read_only-Modus iter_rows() verwenden (viel schneller als cell())
    col_ean = None
    col_product = None
    col_price = None
    col_order = None
    header_found = False
    rows_to_skip = 0

    for row in ws.iter_rows(min_row=1, max_row=5, values_only=True):
        rows_to_skip += 1
        cells_lower = [str(c).lower().strip() if c else "" for c in row]

        if any("ean" in c for c in cells_lower) and any("order" in c or "quantity" in c or "bestellung" in c for c in cells_lower):
            header_found = True
            for col_idx, cell_val in enumerate(cells_lower):
                if "ean" in cell_val and col_ean is None:
                    col_ean = col_idx
                elif ("product" in cell_val or "produkt" in cell_val or "beschreibung" in cell_val or "description" in cell_val):
                    if col_product is None:
                        col_product = col_idx
                elif ("price" in cell_val or "preis" in cell_val) and "bulk" not in cell_val and "120" not in cell_val:
                    if col_price is None:
                        col_price = col_idx
                elif "order" in cell_val or "bestellung" in cell_val or "quantity" in cell_val:
                    if col_order is None:
                        col_order = col_idx
            break

    if not header_found:
        wb.close()
        raise ValueError(
            "Konnte die Spaltenüberschriften nicht erkennen.\n"
            "Die Excel-Datei muss Spalten mit 'EAN' und 'Order' enthalten."
        )

    if not all(x is not None for x in [col_ean, col_product, col_price, col_order]):
        missing = []
        if col_ean is None:
            missing.append("EAN")
        if col_product is None:
            missing.append("Product/Produkt")
        if col_price is None:
            missing.append("Price/Preis")
        if col_order is None:
            missing.append("Order/Bestellung")
        wb.close()
        raise ValueError(f"Fehlende Spalten: {', '.join(missing)}")

    # Datenzeilen einlesen – iter_rows() ist 10-50x schneller als cell()
    # weil openpyxl im read_only-Modus die Zeilen streamt
    items = []
    for row in ws.iter_rows(min_row=rows_to_skip + 1, values_only=True):
        # Schneller Abbruch: Order-Spalte zuerst prüfen
        ncols = len(row)
        if col_order >= ncols:
            continue
        order_val = row[col_order]
        if order_val is None:
            continue
        try:
            order_qty = int(float(order_val))
        except (ValueError, TypeError):
            continue
        if order_qty <= 0:
            continue

        if col_ean >= ncols:
            continue
        ean_raw = row[col_ean]
        if ean_raw is None:
            continue

        # EAN als String formatieren (ohne Dezimalstellen)
        if isinstance(ean_raw, float):
            ean_str = str(int(ean_raw))
        else:
            ean_str = str(ean_raw).strip()

        product_raw = row[col_product] if col_product < ncols else ""
        price_raw = row[col_price] if col_price < ncols else 0

        try:
            price = float(price_raw) if price_raw else 0.0
        except (ValueError, TypeError):
            price = 0.0

        items.append({
            "ean": ean_str,
            "product": shorten_product_name(product_raw or ""),
            "quantity": order_qty,
            "source_price": price,
        })

    wb.close()

    # Alphabetisch nach Produktname sortieren
    items.sort(key=lambda x: x["product"].lower())

    return items

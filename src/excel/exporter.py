"""Export einer Rechnung als Excel-Datei im Layout der PDF-Rechnung.

Bildet dieselben Abschnitte wie InvoiceGenerator (src/pdf/invoice.py) ab:
Firmen- und Kundendaten, Rechnungsnummer/Datum, Positionstabelle sowie die
Netto/USt./Brutto-Zusammenfassung und den Footer-Text. Die Preis- und
Summenspalten bleiben als Formeln editierbar (Menge * Einzelpreis usw.),
sodass manuell angepasste Werte automatisch neu berechnet werden.
"""

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from src.config import COMPANY, FOOTER

_HEADER_FILL = PatternFill("solid", fgColor="D9D9D9")
_THIN        = Side(style="thin", color="000000")
_BORDER_ALL  = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_CURRENCY_FMT = '#,##0.00 "€"'
_BOLD         = Font(bold=True)
_RIGHT        = Alignment(horizontal="right")
_CENTER       = Alignment(horizontal="center")

_TABLE_HEADERS = ("Stk.", "EAN", "Produkt", "Einzelpreis (Netto) €", "Gesamtpreis (Netto) €")
_COL_WIDTHS    = (8, 16, 50, 20, 20)

_LABEL_COL = 4   # Spalte "Einzelpreis" — Summen-Beschriftungen stehen direkt links der Beträge
_VALUE_COL = 5   # Spalte "Gesamtpreis" — Summenbeträge


def _write(ws, row, col, value, *, font=None, alignment=None, number_format=None):
    cell = ws.cell(row=row, column=col, value=value)
    if font:
        cell.font = font
    if alignment:
        cell.alignment = alignment
    if number_format:
        cell.number_format = number_format
    return cell


def _draw_company_block(ws, row):
    _write(ws, row, 1, COMPANY["name"], font=Font(bold=True, size=12))
    row += 1
    for line in (COMPANY["street"], COMPANY["city"], COMPANY["phone"],
                 COMPANY["email"], COMPANY["atu"], COMPANY["eori"]):
        _write(ws, row, 1, line)
        row += 1
    return row


def _draw_customer_block(ws, row, customer):
    _write(ws, row, 1, "An", font=_BOLD)
    row += 1
    _write(ws, row, 1, customer.get("name", ""), font=_BOLD)
    row += 1
    for field in ("street", "plz_city", "country"):
        value = customer.get(field, "").strip()
        if value:
            _write(ws, row, 1, value)
            row += 1
    vat = customer.get("vat", "").strip()
    if vat:
        _write(ws, row, 1, f"VAT: {vat}")
        row += 1
    return row


def _draw_invoice_meta(ws, row, invoice_data):
    _write(ws, row, 1, f"Rechnung Nr. {invoice_data.get('number', '')}", font=Font(bold=True, size=12))
    _write(ws, row, _LABEL_COL, f"Rechnungsdatum: {invoice_data.get('date', '')}")
    row += 1
    if invoice_data.get("is_export", False):
        _write(ws, row, 1, "EXPORT", font=Font(bold=True, size=12), alignment=_CENTER)
        row += 1
    return row


def _draw_table_header(ws, row, ust_enabled):
    headers = list(_TABLE_HEADERS)
    if ust_enabled:
        headers.append("USt. %")
    for col, header in enumerate(headers, start=1):
        _write(ws, row, col, header, font=_BOLD, alignment=_CENTER, number_format=None)
        cell = ws.cell(row=row, column=col)
        cell.fill = _HEADER_FILL
        cell.border = _BORDER_ALL
    return row + 1, len(headers)


def _draw_items(ws, row, items, ust_enabled, ust_percent):
    first_row = row
    total_qty = 0
    for item in items:
        quantity   = item["quantity"]
        unit_price = item["unit_price"]
        total_qty += quantity

        _write(ws, row, 1, quantity, alignment=_CENTER)
        _write(ws, row, 2, item["ean"], alignment=_CENTER)
        _write(ws, row, 3, item["product"])
        _write(ws, row, 4, round(unit_price, 2), number_format=_CURRENCY_FMT)
        _write(ws, row, 5, f"=A{row}*D{row}", number_format=_CURRENCY_FMT)
        if ust_enabled:
            _write(ws, row, 6, ust_percent, alignment=_CENTER, number_format='0"%"')

        last_col = 6 if ust_enabled else 5
        for col in range(1, last_col + 1):
            ws.cell(row=row, column=col).border = _BORDER_ALL
        row += 1

    return row, first_row, row - 1, total_qty


def _draw_summary(ws, row, first_item_row, last_item_row, total_qty, ust_enabled, ust_percent):
    if last_item_row >= first_item_row:
        netto_formula = f"=SUM(E{first_item_row}:E{last_item_row})"
    else:
        netto_formula = 0

    _write(ws, row, 1, total_qty, font=_BOLD, alignment=_CENTER)
    _write(ws, row, _LABEL_COL, "Netto", font=_BOLD, alignment=_RIGHT)
    netto_cell = _write(ws, row, _VALUE_COL, netto_formula, font=_BOLD, number_format=_CURRENCY_FMT)
    netto_row = row
    row += 1

    if ust_enabled:
        _write(ws, row, _LABEL_COL, "USt.", font=_BOLD, alignment=_RIGHT)
        _write(ws, row, _VALUE_COL, f"=E{netto_row}*{ust_percent}/100",
               font=_BOLD, number_format=_CURRENCY_FMT)
        _write(ws, row, 6, f"{ust_percent:.0f}%", alignment=_CENTER)
        ust_row = row
        row += 1

        _write(ws, row, _LABEL_COL, "Gesamtsumme Brutto", font=_BOLD, alignment=_RIGHT)
        _write(ws, row, _VALUE_COL, f"=E{netto_row}+E{ust_row}",
               font=_BOLD, number_format=_CURRENCY_FMT)
    else:
        _write(ws, row, _LABEL_COL, "Gesamtsumme Netto", font=_BOLD, alignment=_RIGHT)
        _write(ws, row, _VALUE_COL, f"=E{netto_row}", font=_BOLD, number_format=_CURRENCY_FMT)

    return row + 1


def _draw_footer(ws, row, invoice_data):
    if invoice_data.get("is_export", False):
        lines = [FOOTER.get("delivery_terms", "")]
    else:
        lines = [FOOTER.get("eu_text_1", ""), FOOTER.get("eu_text_2", ""), FOOTER.get("eu_text_3", "")]
    for line in lines:
        if line:
            _write(ws, row, 1, line)
            row += 1
    return row


def export_items_to_excel(items, invoice_data, customer_data, output_path):
    """Schreibt eine Rechnung als .xlsx im selben Aufbau wie die PDF-Rechnung.

    Args:
        items: Liste von dicts mit ean, product, quantity, unit_price
               (gleiche Struktur wie an generate_invoice() übergeben)
        invoice_data: dict mit number, date, ust_enabled, ust_percent, is_export
        customer_data: dict mit name, street, plz_city, country, vat
        output_path: Pfad für die .xlsx-Datei
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Rechnung"

    ust_enabled = invoice_data.get("ust_enabled", False)
    ust_percent = invoice_data.get("ust_percent", 0) if ust_enabled else 0

    row = _draw_company_block(ws, 1)
    row = _draw_customer_block(ws, row + 1, customer_data)
    row = _draw_invoice_meta(ws, row + 1, invoice_data)

    header_row, num_cols = _draw_table_header(ws, row + 2, ust_enabled)
    row, first_item_row, last_item_row, total_qty = _draw_items(
        ws, header_row, items, ust_enabled, ust_percent)

    row = _draw_summary(ws, row + 1, first_item_row, last_item_row,
                        total_qty, ust_enabled, ust_percent)
    row = _draw_footer(ws, row + 2, invoice_data)

    widths = list(_COL_WIDTHS)
    if ust_enabled:
        widths.append(10)
    for col, width in enumerate(widths[:num_cols], start=1):
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)
    wb.save(output_path)

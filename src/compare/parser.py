"""
Parser für den VergleichsBot: liest Excel- und PDF-Dateien,
erkennt Spalten automatisch per Keywords.
"""

import openpyxl

_QUANTITY_KW = {
    # Deutsch
    "menge", "anzahl", "anz", "stk", "stück", "stückzahl",
    "bestellmenge", "bestellte", "bedarfsmenge", "einheit", "einheiten",
    "karton", "ktn", "krt", "pakete", "pack",
    # Englisch
    "quantity", "qty", "order", "ordered", "amount",
    "pcs", "pieces", "units", "count",
}
_EAN_KW = {
    "ean", "ean13", "ean-13", "ean code", "ean-code",
    "barcode", "bar code", "bar-code",
    "gtin", "gtin-13",
}
_SKU_KW = {
    # Deutsch
    "sku", "artikel", "artikelnummer", "artikelcode", "artikelnr",
    "art.-nr", "art.nr", "art-nr", "art. nr", "art nr",
    "bestell-nr", "bestellnummer", "bestellnr",
    "referenz", "ref.-nr", "ref.nr", "ref",
    "pos", "position",
    # Englisch
    "article", "item", "item no", "item-no", "item#",
    "product no", "product-no", "product#", "prod. no",
    "part", "part no", "part-no", "partno",
    "code",
}
_PRODUCT_KW = {
    # Deutsch
    "produkt", "produktname", "bezeichnung", "beschreibung",
    "warenbezeichnung", "ware", "titel",
    # Englisch
    "product", "product name", "description", "name", "item name", "item description",
}


def parse_file(filepath):
    """
    Liest eine Excel- oder PDF-Datei und gibt Positionen zurück.

    Returns:
        list[dict]: [{ean, sku, product, quantity}]
    Raises:
        ValueError: wenn Spalten nicht erkannt werden oder Datei leer ist
        ImportError: wenn pdfplumber fehlt
    """
    ext = filepath.lower().rsplit(".", 1)[-1]
    if ext in ("xlsx", "xls"):
        return _parse_excel(filepath)
    elif ext == "pdf":
        return _parse_pdf(filepath)
    else:
        raise ValueError(f"Nicht unterstütztes Dateiformat: .{ext}\nBitte Excel (.xlsx) oder PDF verwenden.")


def _detect_columns(row):
    """Gibt (col_ean, col_sku, col_product, col_quantity) zurück — jede kann None sein."""
    col_ean = col_sku = col_product = col_quantity = None
    for i, cell in enumerate(row):
        if not cell:
            continue
        c = str(cell).lower().strip()
        if col_ean is None and any(kw in c for kw in _EAN_KW):
            col_ean = i
        elif col_sku is None and any(kw in c for kw in _SKU_KW):
            col_sku = i
        elif col_product is None and any(kw in c for kw in _PRODUCT_KW):
            col_product = i
        elif col_quantity is None and any(kw in c for kw in _QUANTITY_KW):
            col_quantity = i

    # Fallback: wenn Quantity erkannt aber kein Identifier, nimm die erste nicht-Quantity-Spalte
    if col_quantity is not None and not _has_identifier(col_ean, col_sku, col_product):
        for i, cell in enumerate(row):
            if i != col_quantity and cell and str(cell).strip():
                col_product = i
                break

    return col_ean, col_sku, col_product, col_quantity


def _has_identifier(col_ean, col_sku, col_product):
    return col_ean is not None or col_sku is not None or col_product is not None


def _extract_row(row, col_ean, col_sku, col_product, col_quantity):
    """Liest eine Datenzeile aus. Gibt None zurück wenn Zeile ungültig oder leer."""
    ncols = len(row)
    if col_quantity is None or col_quantity >= ncols:
        return None
    qty_raw = row[col_quantity]
    if qty_raw is None:
        return None
    try:
        qty = int(float(qty_raw))
    except (ValueError, TypeError):
        return None
    if qty <= 0:
        return None

    def _str(col):
        if col is None or col >= ncols:
            return ""
        val = row[col]
        if val is None:
            return ""
        if isinstance(val, float) and val == int(val):
            return str(int(val))
        return str(val).strip()

    ean     = _str(col_ean)
    sku     = _str(col_sku)
    product = _str(col_product)
    if not ean and not sku and not product:
        return None
    return {"ean": ean, "sku": sku, "product": product, "quantity": qty}


def _parse_excel(filepath):
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb.active

    col_ean = col_sku = col_product = col_quantity = None
    header_row_idx = None

    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True)):
        col_ean, col_sku, col_product, col_quantity = _detect_columns(row)
        if col_quantity is not None and _has_identifier(col_ean, col_sku, col_product):
            header_row_idx = row_idx + 1
            break

    if header_row_idx is None:
        wb.close()
        raise ValueError(
            "Konnte keine Spaltenüberschriften erkennen.\n"
            "Die Datei muss eine Mengen-Spalte enthalten (z.B. 'Quantity', 'Menge', 'Stk.', 'Anzahl', 'Qty', 'Pcs')."
        )

    items = []
    for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
        item = _extract_row(row, col_ean, col_sku, col_product, col_quantity)
        if item:
            items.append(item)

    wb.close()

    if not items:
        raise ValueError("Die Datei enthält keine gültigen Positionen mit Menge > 0.")
    return items


def _parse_pdf(filepath):
    try:
        import pdfplumber
    except ImportError:
        raise ImportError(
            "pdfplumber ist nicht installiert.\n"
            "Bitte in der Kommandozeile ausführen:\n"
            "pip install pdfplumber"
        )

    col_ean = col_sku = col_product = col_quantity = None
    header_found = False
    items = []

    with pdfplumber.open(filepath) as pdf:
        for page in pdf.pages:
            tables = page.extract_tables()
            for table in tables:
                if not table:
                    continue
                for row in table:
                    if not header_found:
                        ce, cs, cp, cq = _detect_columns(row)
                        if cq is not None and _has_identifier(ce, cs, cp):
                            col_ean, col_sku, col_product, col_quantity = ce, cs, cp, cq
                            header_found = True
                    else:
                        item = _extract_row(row, col_ean, col_sku, col_product, col_quantity)
                        if item:
                            items.append(item)

    if not header_found:
        raise ValueError(
            "Konnte keine Tabelle mit Spaltenüberschriften im PDF finden.\n"
            "Das PDF muss eine Tabelle mit Mengen-Spalte enthalten (z.B. 'Quantity', 'Menge', 'Stk.', 'Anzahl', 'Qty', 'Pcs')."
        )
    if not items:
        raise ValueError("Das PDF enthält keine gültigen Positionen mit Menge > 0.")
    return items

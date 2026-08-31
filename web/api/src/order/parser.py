"""
Parser für den Order-Tab: liest die Bestellliste (EAN + Order-Menge) ein
und baut ein EAN -> Menge Lookup. Nutzt dieselben Spalten-Keywords wie der
VergleichsBot-Parser.
"""

import openpyxl

from src.compare.parser import _EAN_KW, _QUANTITY_KW


def normalize_ean(value):
    """Normalisiert einen EAN-Zellwert auf einen vergleichbaren String.

    Excel liefert numerische EANs oft als float (z.B. 4001234500000.0) —
    ohne Normalisierung würde das nie gegen einen String-EAN aus einer
    anderen Datei matchen.
    """
    if value is None:
        return ""
    if isinstance(value, float) and value == int(value):
        return str(int(value))
    return str(value).strip()


def _detect_columns(row):
    """Gibt (col_ean, col_qty) zurück — jede kann None sein."""
    col_ean = col_qty = None
    for i, cell in enumerate(row):
        if not cell:
            continue
        c = str(cell).lower().strip()
        if col_ean is None and any(kw in c for kw in _EAN_KW):
            col_ean = i
        elif col_qty is None and any(kw in c for kw in _QUANTITY_KW):
            col_qty = i
    return col_ean, col_qty


def parse_order_source(filepath):
    """
    Liest die Bestellliste ein und summiert die Order-Menge pro EAN.

    Returns:
        dict[str, int]: EAN -> Summe der Bestellmenge
    Raises:
        ValueError: wenn keine EAN-/Order-Spalte erkannt wird oder keine
            gültigen Zeilen vorhanden sind.
    """
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb.active

    col_ean = col_qty = None
    header_row_idx = None
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=20, values_only=True)):
        col_ean, col_qty = _detect_columns(row)
        if col_ean is not None and col_qty is not None:
            header_row_idx = row_idx + 1
            break

    if header_row_idx is None:
        wb.close()
        raise ValueError(
            "Konnte keine EAN- und Order-Spalte in der Bestellliste erkennen.\n"
            "Die Datei muss Spalten für EAN und Bestellmenge enthalten "
            "(z.B. 'EAN', 'Order', 'Menge')."
        )

    lookup = {}
    max_idx = max(col_ean, col_qty)
    for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
        if len(row) <= max_idx:
            continue
        ean = normalize_ean(row[col_ean])
        if not ean:
            continue
        try:
            qty = int(float(row[col_qty]))
        except (TypeError, ValueError):
            continue
        if qty <= 0:
            continue
        lookup[ean] = lookup.get(ean, 0) + qty

    wb.close()

    if not lookup:
        raise ValueError("Die Bestellliste enthält keine gültigen Zeilen mit EAN und Menge > 0.")
    return lookup

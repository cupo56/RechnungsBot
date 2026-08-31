"""
Matcher für den Order-Tab: liest die Ziel-Liste vollständig ein und hängt
eine Order-Spalte an, deren Werte aus dem Bestellliste-Lookup (EAN -> Menge)
stammen.
"""

from dataclasses import dataclass

import openpyxl

from src.compare.parser import _EAN_KW
from src.order.parser import normalize_ean

NOT_FOUND = "Nicht gefunden"


@dataclass
class TargetList:
    """Vollständig eingelesene Ziel-Liste (alle Original-Spalten/-Zeilen)."""

    headers: list
    rows: list
    ean_col: int

    def __len__(self):
        return len(self.rows)


@dataclass
class OrderResult:
    """Ziel-Liste + angehängte Order-Spalte, plus Trefferstatistik."""

    headers: list
    rows: list
    n_matched: int
    n_not_found: int


def read_target_list(filepath):
    """
    Liest die Ziel-Liste vollständig ein (alle Spalten, alle Zeilen) und
    erkennt die EAN-Spalte.

    Returns:
        TargetList
    Raises:
        ValueError: wenn keine EAN-Spalte erkannt wird oder keine
            Datenzeilen vorhanden sind.
    """
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb.active

    header_row_idx = None
    col_ean = None
    header_values = None
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=20, values_only=True)):
        for i, cell in enumerate(row):
            if cell and any(kw in str(cell).lower().strip() for kw in _EAN_KW):
                col_ean = i
                break
        if col_ean is not None:
            header_row_idx = row_idx + 1
            header_values = row
            break

    if header_row_idx is None:
        wb.close()
        raise ValueError(
            "Konnte keine EAN-Spalte in der Ziel-Liste erkennen.\n"
            "Die Datei muss eine Spalte für EAN enthalten (z.B. 'EAN', 'Barcode', 'GTIN')."
        )

    headers = [str(h).strip() if h is not None else "" for h in header_values]
    rows = []
    for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
        if all(cell is None for cell in row):
            continue
        rows.append(list(row))

    wb.close()

    if not rows:
        raise ValueError("Die Ziel-Liste enthält keine Datenzeilen.")

    return TargetList(headers=headers, rows=rows, ean_col=col_ean)


def apply_order_column(target, order_lookup):
    """
    Hängt an jede Zeile der Ziel-Liste die nachgeschlagene Order-Menge an.

    Args:
        target: TargetList (von read_target_list())
        order_lookup: dict[str, int] (von parse_order_source())

    Returns:
        OrderResult
    """
    headers = target.headers + ["Order"]
    ncols = len(target.headers)
    rows = []
    n_matched = 0
    n_not_found = 0

    for row in target.rows:
        padded = list(row) + [None] * (ncols - len(row))
        ean = normalize_ean(padded[target.ean_col])
        qty = order_lookup.get(ean)
        if qty is not None:
            n_matched += 1
            rows.append(padded + [qty])
        else:
            n_not_found += 1
            rows.append(padded + [NOT_FOUND])

    return OrderResult(headers=headers, rows=rows, n_matched=n_matched, n_not_found=n_not_found)

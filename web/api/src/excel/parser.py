"""
Excel-Parser für RechnungsBot.
Liest Excel-Dateien verschiedener Lieferanten ein und extrahiert bestellte
Positionen.  Unterstützt .xlsx (openpyxl) und optional .xls (xlrd).

Erkennt Spaltenköpfe anhand erweiterter Keyword-Sets (DE/EN/CZ/PL),
durchsucht alle Tabellenblätter und bis zu 20 Zeilen pro Blatt.
"""

import os
import re

import openpyxl

# Optionaler .xls-Support ─────────────────────────────────────────────
try:
    import xlrd  # type: ignore

    _HAS_XLRD = True
except ImportError:
    _HAS_XLRD = False

# ─── Regex-Patterns ──────────────────────────────────────────────────
_RE_GENDER = re.compile(r'\s*\((man|woman|unisex)\)\s*', re.IGNORECASE)
_RE_COVER = re.compile(
    r'\s*\((New Cover|Old Cover|Cover with [^)]+|Without box|[A-Za-z]+ Cover[^)]*)\)\s*',
    re.IGNORECASE,
)
_RE_WHITESPACE = re.compile(r'\s{2,}')
_RE_CURRENCY = re.compile(r'[€$£¥₹₽\s]')

# Parfüm-Bezeichnungen: längere Varianten zuerst, damit kein Teilmatch vorgreift
_FRAGRANCE_REPLACEMENTS = [
    (re.compile(r'Eau De Parfum Intense', re.IGNORECASE), 'EdP Intense'),
    (re.compile(r'Eau De Parfum', re.IGNORECASE), 'EdP'),
    (re.compile(r'Eau De Toilette', re.IGNORECASE), 'EdT'),
    (re.compile(r'Eau De Cologne Intense', re.IGNORECASE), 'EdC Intense'),
    (re.compile(r'Eau De Cologne', re.IGNORECASE), 'EdC'),
    (re.compile(r'Extrait de Parfum', re.IGNORECASE), 'Extrait'),
    (re.compile(r'Parfum Intense', re.IGNORECASE), 'Parfum Intense'),
]

# ─── Keyword-Sets für Spaltenköpfe ───────────────────────────────────
# "amount" ist absichtlich NICHT enthalten: in diesem Projekt (siehe
# ALINA-PDF-Format in src/pdf_input/parser.py) bezeichnet "Amount" die
# Gesamtsumme einer Zeile, nicht die Bestellmenge – als Mengen-Keyword
# würde es z.B. eine "Total Amount"-Spalte fälschlich als Menge erkennen.
_QUANTITY_KEYWORDS = {
    "order", "quantity", "bestellung", "qty", "menge",
    "stück", "stk", "anzahl", "pcs",
    "ilość", "množství",
}

_PRODUCT_KEYWORDS = {
    "product", "produkt", "beschreibung", "description",
    "název", "artikelbezeichnung", "artikel", "nazwa",
    "bezeichnung", "item", "article", "name", "towar",
}

_PRICE_KEYWORDS = {
    "price", "preis", "cena", "einzelpreis", "unit price",
    "e.k.", "ek", "vk", "netto", "unitprice",
}

# Spalten mit diesen Keywords werden NICHT als Preis erkannt
# ("gewicht" verhindert, dass "netto" fälschlich in "Nettogewicht" matcht)
_PRICE_EXCLUDE = {"bulk", "120", "total", "gesamt", "brutto", "summe", "gewicht"}

_EAN_KEYWORDS = {"ean", "barcode", "gtin", "upc"}

# Maximale Zeile bis zu der nach Spaltenköpfen gesucht wird
_MAX_HEADER_SEARCH_ROW = 20


# ─── Hilfsfunktionen ─────────────────────────────────────────────────

def _matches_any(cell_val: str, keywords: set[str]) -> bool:
    """Prüft ob *irgendein* Keyword als Substring im Zellenwert vorkommt."""
    return any(kw in cell_val for kw in keywords)


def shorten_product_name(name):
    """
    Kürzt Produktbeschreibungen für die Rechnung.
    z.B. 'Boss Boss Bottled Eau De Toilette 50 ml (man)' → 'Boss Boss Bottled EdT 50 ml'
    """
    if not name:
        return ""

    text = str(name).strip()
    text = _RE_GENDER.sub(' ', text)
    text = _RE_COVER.sub(' ', text)

    for pattern, replacement in _FRAGRANCE_REPLACEMENTS:
        text = pattern.sub(replacement, text)

    # Großgeschriebene Abkürzungen normalisieren
    text = text.replace(' EDT ', ' EdT ').replace(' EDP ', ' EdP ').replace(' EDC ', ' EdC ')
    if text.endswith(' EDT'):
        text = text[:-4] + ' EdT'
    if text.endswith(' EDP'):
        text = text[:-4] + ' EdP'

    return _RE_WHITESPACE.sub(' ', text).strip()


def _parse_price(value) -> float:
    """Robuste Konvertierung eines Zellwerts in einen float-Preis.

    Behandelt:
    - Numerische Werte (int/float)
    - Strings mit Währungssymbolen (€, $, £, …)
    - Deutsches Zahlenformat (1.234,56)
    - Englisches Zahlenformat (1,234.56)
    """
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)

    text = _RE_CURRENCY.sub('', str(value).strip())
    if not text:
        return 0.0

    # Deutsches vs. englisches Format
    if ',' in text and '.' in text:
        if text.rindex(',') > text.rindex('.'):
            # 1.234,56 → deutsches Format
            text = text.replace('.', '').replace(',', '.')
        else:
            # 1,234.56 → englisches Format
            text = text.replace(',', '')
    elif ',' in text:
        # Nur Komma → deutsches Dezimaltrennzeichen
        text = text.replace(',', '.')

    try:
        return float(text)
    except (ValueError, TypeError):
        return 0.0


def _find_header(row_iterator, max_rows: int = _MAX_HEADER_SEARCH_ROW):
    """Durchsucht Zeilen nach einer Kopfzeile mit Mengen-Keyword.

    Args:
        row_iterator: Iterator über Zeilen (Tupel von Zellwerten).
        max_rows: Maximale Anzahl zu durchsuchender Zeilen.

    Returns:
        tuple: (header_row_number, col_map) oder (None, None).
               col_map hat Schlüssel 'ean', 'product', 'price', 'order'.
    """
    for row_num, row in enumerate(row_iterator, start=1):
        if row_num > max_rows:
            break

        cells_lower = [str(c).lower().strip() if c else "" for c in row]

        # Enthält die Zeile ein Mengen-/Bestellungs-Keyword?
        if not any(_matches_any(c, _QUANTITY_KEYWORDS) for c in cells_lower):
            continue

        # Spalten zuordnen
        col_map = {'ean': None, 'product': None, 'price': None, 'order': None}

        for col_idx, cell_val in enumerate(cells_lower):
            if not cell_val:
                continue

            if _matches_any(cell_val, _EAN_KEYWORDS) and col_map['ean'] is None:
                col_map['ean'] = col_idx
            elif _matches_any(cell_val, _PRODUCT_KEYWORDS) and col_map['product'] is None:
                col_map['product'] = col_idx
            elif (
                _matches_any(cell_val, _PRICE_KEYWORDS)
                and not _matches_any(cell_val, _PRICE_EXCLUDE)
                and col_map['price'] is None
            ):
                col_map['price'] = col_idx
            elif _matches_any(cell_val, _QUANTITY_KEYWORDS) and col_map['order'] is None:
                col_map['order'] = col_idx

        # Nur akzeptieren, wenn alle Pflichtspalten gefunden wurden – sonst war
        # es ein falsches Positiv (z.B. eine Titelzeile mit "Bestellung" im
        # Text, ohne echte Produkt-/Preis-Spalten) und wir suchen in einer
        # der folgenden Zeilen weiter, statt das Blatt aufzugeben.
        if (
            col_map['product'] is not None
            and col_map['price'] is not None
            and col_map['order'] is not None
        ):
            return row_num, col_map

    return None, None


def _extract_items(row_iterator, col_map: dict) -> list[dict]:
    """Extrahiert Positionen aus Datenzeilen anhand der Spaltenzuordnung."""
    col_ean = col_map['ean']
    col_product = col_map['product']
    col_price = col_map['price']
    col_order = col_map['order']

    items = []
    for row in row_iterator:
        ncols = len(row)

        # Schneller Abbruch: Order-Spalte zuerst prüfen
        if col_order >= ncols:
            continue
        order_val = row[col_order]
        if order_val is None:
            continue
        try:
            order_qty = int(float(order_val))
        except (ValueError, TypeError):
            continue
        if order_qty <= 0:
            continue

        # EAN als String formatieren (ohne Dezimalstellen) – leer wenn Spalte fehlt
        if col_ean is not None and col_ean < ncols:
            ean_raw = row[col_ean]
            if ean_raw is None:
                ean_str = ""
            elif isinstance(ean_raw, float):
                ean_str = str(int(ean_raw))
            else:
                ean_str = str(ean_raw).strip()
        else:
            ean_str = ""

        product_raw = row[col_product] if col_product is not None and col_product < ncols else ""
        price_raw = row[col_price] if col_price is not None and col_price < ncols else 0

        items.append({
            "ean": ean_str,
            "product": shorten_product_name(product_raw or ""),
            "quantity": order_qty,
            "source_price": _parse_price(price_raw),
        })

    return items


def _build_error_message(searched_sheets: list[str]) -> str:
    """Erzeugt eine strukturierte Fehlermeldung mit Diagnose-Hinweisen."""
    qty_kw = ", ".join(sorted(_QUANTITY_KEYWORDS))
    prod_kw = ", ".join(sorted(_PRODUCT_KEYWORDS))
    price_kw = ", ".join(sorted(_PRICE_KEYWORDS))
    ean_kw = ", ".join(sorted(_EAN_KEYWORDS))

    return (
        "Konnte keine Spaltenüberschriften erkennen.\n"
        "\n"
        f"Durchsuchte Blätter: {', '.join(searched_sheets) if searched_sheets else '(keine)'}\n"
        f"Durchsuchte Zeilen: 1–{_MAX_HEADER_SEARCH_ROW} (pro Blatt)\n"
        "\n"
        "Unterstützte Spaltenbezeichnungen:\n"
        f"  • Menge/Bestellung: {qty_kw}\n"
        f"  • Produkt:          {prod_kw}\n"
        f"  • Preis:            {price_kw}\n"
        f"  • EAN (optional):   {ean_kw}\n"
        "\n"
        "Bitte prüfe, ob die Excel-Datei eine Kopfzeile mit diesen "
        "Bezeichnungen enthält (in den ersten 20 Zeilen)."
    )


# ─── .xls-Support (Legacy) ──────────────────────────────────────────

def _parse_xls_legacy(filepath: str) -> list[dict]:
    """Parst eine .xls-Datei via xlrd. Gibt dieselbe Struktur wie parse_excel() zurück."""
    if not _HAS_XLRD:
        raise ValueError(
            "Die Datei hat das ältere .xls-Format.\n"
            "Bitte konvertiere sie in .xlsx (Excel → Speichern unter → .xlsx) "
            "oder installiere das Python-Paket 'xlrd' (pip install xlrd)."
        )

    wb = xlrd.open_workbook(filepath)
    searched_sheets: list[str] = []

    for sheet in wb.sheets():
        searched_sheets.append(sheet.name)

        # Zeilen als Tupel-Iterator bereitstellen (wie openpyxl values_only)
        def _row_iter(sh=sheet):
            for rx in range(sh.nrows):
                yield tuple(sh.cell_value(rx, cx) for cx in range(sh.ncols))

        rows = _row_iter()
        header_row, col_map = _find_header(rows)

        if header_row is None:
            continue

        # Pflichtfelder prüfen
        required = {"Product/Produkt": col_map['product'], "Price/Preis": col_map['price'],
                     "Order/Bestellung": col_map['order']}
        missing = [name for name, idx in required.items() if idx is None]
        if missing:
            continue

        # Datenzeilen ab header_row+1 lesen
        def _data_iter(sh=sheet, start=header_row):
            for rx in range(start, sh.nrows):
                yield tuple(sh.cell_value(rx, cx) for cx in range(sh.ncols))

        items = _extract_items(_data_iter(), col_map)
        if items:
            items.sort(key=lambda x: x["product"].lower())
            return items

    raise ValueError(_build_error_message(searched_sheets))


# ─── Hauptfunktion ───────────────────────────────────────────────────

def parse_excel(filepath):
    """
    Liest eine Excel-Datei (xlsx oder xls) und gibt bestellte Positionen zurück.

    Durchsucht alle Tabellenblätter und bis zu 20 Zeilen pro Blatt nach
    einer erkennbaren Kopfzeile. Unterstützt Spaltenbezeichnungen in
    Deutsch, Englisch, Tschechisch und Polnisch.

    Returns:
        list[dict]: Liste von Positionen mit Schlüsseln:
            - ean (str): EAN-Code (leer wenn nicht vorhanden)
            - product (str): Gekürzter Produktname
            - quantity (int): Bestellmenge
            - source_price (float): Einkaufspreis aus der Quelldatei
    """
    ext = os.path.splitext(filepath)[1].lower()

    # .xls → Legacy-Pfad
    if ext == '.xls':
        return _parse_xls_legacy(filepath)

    # .xlsx → openpyxl
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    searched_sheets: list[str] = []

    try:
        for ws in wb.worksheets:
            searched_sheets.append(ws.title)

            # Header suchen (Zeile 1–20)
            header_row, col_map = _find_header(
                ws.iter_rows(min_row=1, max_row=_MAX_HEADER_SEARCH_ROW, values_only=True)
            )

            if header_row is None:
                continue

            # Pflichtfelder prüfen
            required = {
                "Product/Produkt": col_map['product'],
                "Price/Preis": col_map['price'],
                "Order/Bestellung": col_map['order'],
            }
            missing = [name for name, idx in required.items() if idx is None]
            if missing:
                # Dieses Blatt hat zwar ein Mengen-Keyword, aber nicht alle Pflichtspalten
                # → nächstes Blatt probieren
                continue

            # Datenzeilen einlesen ab der Zeile nach dem Header
            items = _extract_items(
                ws.iter_rows(min_row=header_row + 1, values_only=True),
                col_map,
            )

            if items:
                items.sort(key=lambda x: x["product"].lower())
                return items

    finally:
        wb.close()

    raise ValueError(_build_error_message(searched_sheets))
